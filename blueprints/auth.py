# 认证 Blueprint
from datetime import datetime
from io import BytesIO
import os
import tempfile

from flask import Blueprint, flash, redirect, render_template, request, send_file, url_for
from flask_login import current_user, login_required, login_user, logout_user
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from werkzeug.security import check_password_hash

from config import DEVICE_STATUS_LABELS
from database import get_db
from utils.audit import log_action

auth_bp = Blueprint("auth", __name__)


@auth_bp.route("/login", methods=["GET", "POST"])
def login():
    """用户登录"""
    if request.method == "POST":
        username = request.form.get("username", "").strip()
        password = request.form.get("password", "").strip()
        next_page = request.form.get("next", "").strip()
        conn = get_db()
        cur = conn.cursor()
        cur.execute("SELECT * FROM users WHERE username = %s", (username,))
        row = cur.fetchone()
        conn.close()
        if row and row["status"] != "active":
            flash("账户已停用，请联系管理员。", "warning")
        elif row and check_password_hash(row["password"], password):
            from models.user import User

            user = User(row["id"], row["username"], row["role"], row["password"])
            login_user(user)
            log_action(user.username, "login", "user", user.id, "用户登录")
            if next_page and next_page.startswith("/"):
                return redirect(next_page)
            return redirect(url_for("dashboard.dashboard"))
        flash("用户名或密码错误。", "danger")
    return render_template("login.html", next=request.args.get("next", ""))


@auth_bp.route("/logout")
@login_required
def logout():
    """用户登出"""
    log_action(current_user.username, "logout", "user", current_user.id, "用户登出")
    logout_user()
    return redirect(url_for("auth.login"))


@auth_bp.route("/devices")
@login_required
def index():
    """设备列表首页"""
    query = request.args.get("q", "").strip()
    show_inactive = request.args.get("show_inactive", "").lower() in {"1", "true", "on", "yes"}
    conn = get_db()
    cur = conn.cursor()
    status_filter = "" if (show_inactive and current_user.is_admin) else "AND status = 'active'"
    if query:
        cur.execute(
            f"""
            SELECT * FROM devices
            WHERE (device_code LIKE %s OR device_name LIKE %s OR model LIKE %s)
            {status_filter}
            ORDER BY created_at DESC
            """,
            (f"%{query}%", f"%{query}%", f"%{query}%"),
        )
    else:
        cur.execute(f"SELECT * FROM devices WHERE 1=1 {status_filter} ORDER BY created_at DESC")
    devices = cur.fetchall()
    conn.close()
    return render_template(
        "index.html", devices=devices, query=query, show_inactive=show_inactive,
    )


def _create_excel_response(wb, filename):
    """将 Workbook 转为 Flask send_file 响应"""
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return send_file(
        output,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name=filename,
    )


def _save_excel_to_temp(wb, filename):
    """将 Workbook 保存到临时目录，返回文件路径"""
    temp_dir = os.path.join(tempfile.gettempdir(), 'DMS_Exports')
    os.makedirs(temp_dir, exist_ok=True)
    filepath = os.path.join(temp_dir, filename)
    wb.save(filepath)
    return filepath


def _set_header_style(cell):
    """设置表头样式"""
    cell.font = Font(bold=True, color="FFFFFF")
    cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    cell.alignment = Alignment(horizontal="center", vertical="center")


def _set_border(cell):
    """设置单元格边框"""
    thin = Side(style="thin", color="000000")
    cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)


@auth_bp.route("/devices/export")
@login_required
def export_devices():
    """导出设备列表为 Excel"""
    query = request.args.get("q", "").strip()
    show_inactive = request.args.get("show_inactive", "").lower() in {"1", "true", "on", "yes"}
    conn = get_db()
    cur = conn.cursor()
    status_filter = "" if (show_inactive and current_user.is_admin) else "AND status = 'active'"
    if query:
        cur.execute(
            f"""
            SELECT * FROM devices
            WHERE (device_code LIKE %s OR device_name LIKE %s OR model LIKE %s)
            {status_filter}
            ORDER BY created_at DESC
            """,
            (f"%{query}%", f"%{query}%", f"%{query}%"),
        )
    else:
        cur.execute(f"SELECT * FROM devices WHERE 1=1 {status_filter} ORDER BY created_at DESC")
    devices = cur.fetchall()
    conn.close()

    wb = Workbook()
    ws = wb.active
    ws.title = "设备列表"

    headers = ["设备编码", "设备名称", "型号", "位置", "状态"]
    ws.append(headers)
    for cell in ws[1]:
        _set_header_style(cell)
        _set_border(cell)

    for device in devices:
        ws.append([
            device["device_code"] or "-",
            device["device_name"] or "-",
            device["model"] or "-",
            device["location"] or "-",
            DEVICE_STATUS_LABELS.get(device["status"], device["status"]),
        ])

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        for cell in row:
            _set_border(cell)
            cell.alignment = Alignment(vertical="center", wrap_text=True)

    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except Exception:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column].width = adjusted_width

    filename = f"devices_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    is_desktop = request.headers.get('X-Desktop-Shell') == '1'
    if is_desktop:
        filepath = _save_excel_to_temp(wb, filename)
        return {"success": True, "filepath": filepath, "filename": filename}
    return _create_excel_response(wb, filename)
